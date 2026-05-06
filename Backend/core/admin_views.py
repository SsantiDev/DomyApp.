import logging
import io
import openpyxl
from openpyxl.styles import Font as XLFont, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import timedelta
from django.utils import timezone
from django.db import models, transaction
from django.db.models import Sum, Count
from django.conf import settings
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication
from apps.services.models import ServiceRequest

# ReportLab imports for PDF generation
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

logger = logging.getLogger(__name__)


def require_admin(request):
    return request.user.is_authenticated and request.user.role == 'ADMIN'


class FinanceSummaryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not require_admin(request):
            return Response({'detail': 'Forbidden.'}, status=status.HTTP_403_FORBIDDEN)

        # Date filtering params
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        
        queryset = ServiceRequest.objects.filter(status=ServiceRequest.Status.COMPLETED)
        
        if start_date_str:
            try:
                start_date = timezone.datetime.strptime(start_date_str, '%Y-%m-%d')
                queryset = queryset.filter(completed_at__gte=start_date)
            except ValueError:
                pass
                
        if end_date_str:
            try:
                end_date = timezone.datetime.strptime(end_date_str, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
                queryset = queryset.filter(completed_at__lte=end_date)
            except ValueError:
                pass

        total_revenue = queryset.aggregate(total=Sum('total_price'))['total'] or 0
        platform_commissions = float(total_revenue) * settings.PLATFORM_FEE_PERCENTAGE
        
        # Calculate pending payouts: Completed requests that are not yet billed/payout completed.
        # Operators get 80% (1 - PLATFORM_FEE_PERCENTAGE)
        pending_payouts_raw = queryset.filter(is_billed=False).aggregate(total=Sum('total_price'))['total'] or 0
        pending_payouts = float(pending_payouts_raw) * (1 - settings.PLATFORM_FEE_PERCENTAGE)

        # Group by month chronologically for the last 6 months (or for custom date range)
        monthly_data = {}
        now = timezone.now()
        month_names = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
        
        # Initialize last 6 months with 0
        for i in range(5, -1, -1):
            target_date = now - timedelta(days=i*30)
            m_idx = target_date.month - 1
            year = target_date.year
            key = f"{month_names[m_idx]} {year}"
            monthly_data[key] = {
                'month': month_names[m_idx],
                'year': year,
                'value': 0.0,
                'key': key
            }
            
        # Accumulate values from the queryset
        for s in queryset:
            if s.completed_at:
                m_idx = s.completed_at.month - 1
                year = s.completed_at.year
                key = f"{month_names[m_idx]} {year}"
                
                if key not in monthly_data:
                    monthly_data[key] = {
                        'month': month_names[m_idx],
                        'year': year,
                        'value': 0.0,
                        'key': key
                    }
                monthly_data[key]['value'] += float(s.total_price)
                
        # Sort monthly data chronologically
        sorted_months = sorted(
            monthly_data.values(),
            key=lambda x: (x['year'], month_names.index(x['month']))
        )

        return Response({
            'total_revenue': float(total_revenue),
            'platform_commissions': round(platform_commissions, 2),
            'pending_payouts': round(pending_payouts, 2),
            'period': 'custom' if (start_date_str or end_date_str) else 'monthly',
            'monthly_revenues': sorted_months
        })


class WorkerPenaltyView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, worker_id):
        if not require_admin(request):
            return Response({'detail': 'Forbidden.'}, status=status.HTTP_403_FORBIDDEN)

        from apps.users.models import User
        try:
            worker = User.objects.get(pk=worker_id, role='WORKER')
        except User.DoesNotExist:
            return Response({'detail': 'Worker not found.'}, status=status.HTTP_404_NOT_FOUND)

        reason = request.data.get('reason', '')
        deduction = float(request.data.get('deduction', 0.5))

        if hasattr(worker, 'worker_info') and worker.worker_info.average_rating is not None:
            worker.worker_info.average_rating = max(0, float(worker.worker_info.average_rating) - deduction)
            worker.worker_info.save()

        logger.info(
            'Penalty applied to worker %s by admin %s. Reason: %s. Deduction: %s',
            worker_id, request.user.email, reason, deduction
        )
        return Response({
            'status': 'penalty_applied',
            'worker_id': worker_id,
            'reason': reason,
            'applied_by': request.user.email,
        })


class WorkerRejectionRateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not require_admin(request):
            return Response({'detail': 'Forbidden.'}, status=status.HTTP_403_FORBIDDEN)

        since = timezone.now() - timedelta(days=30)
        results = (
            ServiceRequest.objects
            .filter(status=ServiceRequest.Status.CANCELLED, updated_at__gte=since, worker__isnull=False)
            .values('worker__id', 'worker__first_name', 'worker__last_name', 'worker__email')
            .annotate(cancellations=Count('id'))
            .filter(cancellations__gt=0)
            .order_by('-cancellations')
        )
        data = [
            {
                'worker_id': r['worker__id'],
                'worker_name': f"{r['worker__first_name']} {r['worker__last_name']}".strip() or r['worker__email'],
                'cancellations': r['cancellations'],
            }
            for r in results
        ]
        return Response(data)


class FinanceExcelExportView(APIView):
    authentication_classes = []
    permission_classes = []

    def get(self, request):
        # 1. Authenticate via simplejwt manually (supporting both Header and Query Parameter)
        jwt_auth = JWTAuthentication()
        user = None
        
        # Try header authentication first
        try:
            header = jwt_auth.get_header(request)
            if header:
                raw_token = jwt_auth.get_raw_token(header)
                validated_token = jwt_auth.get_validated_token(raw_token)
                user = jwt_auth.get_user(validated_token)
        except Exception:
            pass
            
        # Try query parameter if header wasn't present or failed
        if not user:
            token = request.query_params.get('token')
            if token:
                try:
                    validated_token = jwt_auth.get_validated_token(token)
                    user = jwt_auth.get_user(validated_token)
                except Exception:
                    pass

        if not user or user.role != 'ADMIN':
            return HttpResponse('No autorizado', status=403)

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Finanzas"

        # Apply grid lines visibility
        ws.views.sheetView[0].showGridLines = True

        # Styles
        title_font = XLFont(name='Arial', size=16, bold=True, color='FFFFFF')
        header_font = XLFont(name='Arial', size=11, bold=True, color='FFFFFF')
        data_font = XLFont(name='Arial', size=10)
        total_font = XLFont(name='Arial', size=11, bold=True)
        
        title_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid') # Primary banner
        header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid') # Headers
        zebra_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        total_fill = PatternFill(start_color='EEF2F6', end_color='EEF2F6', fill_type='solid')
        
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        double_bottom_border = Border(
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='double', color='000000')
        )

        # Title block
        ws.merge_cells('A1:I2')
        title_cell = ws['A1']
        title_cell.value = "DOMYAPP - REPORTE DE FINANZAS"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Empty row spacer
        ws.append([])
        ws.append([])

        # Table headers
        headers = [
            "ID Servicio", "Fecha Completado", "Categoría", 
            "Cliente", "Cliente Email", "Operaria", "Operaria Email", 
            "Precio Total", "Comisión Plataforma (20%)"
        ]
        ws.append(headers)
        
        header_row_idx = 4
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row_idx, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        ws.row_dimensions[header_row_idx].height = 25

        # Date filtering params
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        
        services = ServiceRequest.objects.filter(
            status=ServiceRequest.Status.COMPLETED
        ).select_related('client', 'worker', 'category').order_by('-completed_at')
        
        if start_date_str:
            try:
                start_date = timezone.datetime.strptime(start_date_str, '%Y-%m-%d')
                services = services.filter(completed_at__gte=start_date)
            except ValueError:
                pass
                
        if end_date_str:
            try:
                end_date = timezone.datetime.strptime(end_date_str, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
                services = services.filter(completed_at__lte=end_date)
            except ValueError:
                pass

        total_price_sum = 0
        total_commission_sum = 0
        
        current_row = 5
        for i, s in enumerate(services):
            client_name = "N/A"
            if s.client:
                if hasattr(s.client, 'profile'):
                    client_name = f"{s.client.profile.first_name} {s.client.profile.last_name}".strip()
                if not client_name:
                    client_name = s.client.email
                    
            worker_name = "N/A"
            if s.worker:
                if hasattr(s.worker, 'profile'):
                    worker_name = f"{s.worker.profile.first_name} {s.worker.profile.last_name}".strip()
                if not worker_name:
                    worker_name = s.worker.email
                    
            completed_str = s.completed_at.strftime('%Y-%m-%d %H:%M') if s.completed_at else "N/A"
            category_name = s.category.name if s.category else "N/A"
            
            price = float(s.total_price)
            commission = price * settings.PLATFORM_FEE_PERCENTAGE
            
            total_price_sum += price
            total_commission_sum += commission
            
            row_data = [
                s.id, completed_str, category_name,
                client_name, s.client.email if s.client else "",
                worker_name, s.worker.email if s.worker else "",
                price, commission
            ]
            ws.append(row_data)
            
            # Format and borders for data row
            is_even = i % 2 == 1
            for col_num in range(1, 10):
                cell = ws.cell(row=current_row, column=col_num)
                cell.font = data_font
                cell.border = thin_border
                
                # Apply zebra striping
                if is_even:
                    cell.fill = zebra_fill
                
                # Alignments and number formats
                if col_num in [1, 2]:
                    cell.alignment = Alignment(horizontal='center')
                elif col_num in [8, 9]:
                    cell.alignment = Alignment(horizontal='right')
                    cell.number_format = '$#,##0.00'
                else:
                    cell.alignment = Alignment(horizontal='left')
                    
            current_row += 1

        # Totals row
        ws.append([
            "TOTALES", "", "", "", "", "", "",
            total_price_sum, total_commission_sum
        ])
        
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        totals_label_cell = ws.cell(row=current_row, column=1)
        totals_label_cell.font = total_font
        totals_label_cell.alignment = Alignment(horizontal='right', vertical='center')
        
        for col_num in range(1, 10):
            cell = ws.cell(row=current_row, column=col_num)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = double_bottom_border
            if col_num in [8, 9]:
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '$#,##0.00'

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            # Skip merged rows for width calculations
            for cell in col:
                if cell.row > 2 and cell.value:
                    val_str = str(cell.value)
                    # Add extra padding for formatted numbers
                    if cell.number_format == '$#,##0.00':
                        val_str = f"${val_str}.00"
                    max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Prepare HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="reporte_finanzas.xlsx"'
        wb.save(response)
        return response


class FinancePDFExportView(APIView):
    authentication_classes = []
    permission_classes = []

    def get(self, request):
        # 1. Authenticate via simplejwt manually
        jwt_auth = JWTAuthentication()
        user = None
        
        try:
            header = jwt_auth.get_header(request)
            if header:
                raw_token = jwt_auth.get_raw_token(header)
                validated_token = jwt_auth.get_validated_token(raw_token)
                user = jwt_auth.get_user(validated_token)
        except Exception:
            pass
            
        if not user:
            token = request.query_params.get('token')
            if token:
                try:
                    validated_token = jwt_auth.get_validated_token(token)
                    user = jwt_auth.get_user(validated_token)
                except Exception:
                    pass

        if not user or user.role != 'ADMIN':
            return HttpResponse('No autorizado', status=403)

        # Setup PDF document in memory
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )
        
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'DocTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=22,
            textColor=colors.HexColor('#4F46E5'), # Indigo Primary
            spaceAfter=6
        )
        
        subtitle_style = ParagraphStyle(
            'DocSubtitle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            textColor=colors.HexColor('#6B7280'), # Gray
            spaceAfter=20
        )
        
        section_heading = ParagraphStyle(
            'SectionHeading',
            parent=styles['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=14,
            textColor=colors.HexColor('#111827'),
            spaceBefore=10,
            spaceAfter=10
        )
        
        table_header_style = ParagraphStyle(
            'TableHeader',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=8,
            textColor=colors.white,
            alignment=1 # Center
        )
        
        table_cell_style = ParagraphStyle(
            'TableCell',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8,
            textColor=colors.HexColor('#1F2937')
        )
        
        table_cell_center = ParagraphStyle(
            'TableCellCenter',
            parent=table_cell_style,
            alignment=1 # Center
        )
        
        table_cell_right = ParagraphStyle(
            'TableCellRight',
            parent=table_cell_style,
            alignment=2 # Right
        )
        
        metric_label_style = ParagraphStyle(
            'MetricLabel',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=10,
            textColor=colors.HexColor('#4B5563')
        )
        
        metric_value_style = ParagraphStyle(
            'MetricValue',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=16,
            textColor=colors.HexColor('#111827')
        )

        # Header Block
        story.append(Paragraph("DomyApp — Reporte de Finanzas", title_style))
        current_date = timezone.now().strftime('%d de %B de %Y, %H:%M')
        story.append(Paragraph(f"Generado el {current_date} | Administrador: {user.email}", subtitle_style))
        story.append(Spacer(1, 10))

        # Date filtering params
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        
        services = ServiceRequest.objects.filter(
            status=ServiceRequest.Status.COMPLETED
        ).select_related('client', 'worker', 'category').order_by('-completed_at')
        
        if start_date_str:
            try:
                start_date = timezone.datetime.strptime(start_date_str, '%Y-%m-%d')
                services = services.filter(completed_at__gte=start_date)
            except ValueError:
                pass
                
        if end_date_str:
            try:
                end_date = timezone.datetime.strptime(end_date_str, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
                services = services.filter(completed_at__lte=end_date)
            except ValueError:
                pass
        
        total_revenue = sum(float(s.total_price) for s in services)
        platform_commissions = total_revenue * settings.PLATFORM_FEE_PERCENTAGE
        payouts = total_revenue - platform_commissions

        # Render Metrics Box (Summary table)
        metrics_data = [
            [
                Paragraph("Ingresos Totales", metric_label_style),
                Paragraph("Comisión Plataforma (20%)", metric_label_style),
                Paragraph("Pagos a Operarias (80%)", metric_label_style)
            ],
            [
                Paragraph(f"${total_revenue:,.2f}", metric_value_style),
                Paragraph(f"${platform_commissions:,.2f}", metric_value_style),
                Paragraph(f"${payouts:,.2f}", metric_value_style)
            ]
        ]
        
        metrics_table = Table(metrics_data, colWidths=[180, 180, 180])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TEXTCOLOR', (0,0), (-1,-1), colors.HexColor('#111827')),
            ('PADDING', (0,0), (-1,-1), 12),
            ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#E2E8F0')),
            ('LINEBELOW', (0,0), (-1,0), 0.5, colors.HexColor('#E2E8F0')),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0'))
        ]))
        story.append(metrics_table)
        story.append(Spacer(1, 25))

        # Detail Section Title
        story.append(Paragraph("Detalle de Servicios Completados", section_heading))

        # Table for service request details
        headers_def = [
            Paragraph("ID", table_header_style),
            Paragraph("Fecha", table_header_style),
            Paragraph("Categoría", table_header_style),
            Paragraph("Cliente", table_header_style),
            Paragraph("Operaria", table_header_style),
            Paragraph("Total", table_header_style),
            Paragraph("Comisión", table_header_style)
        ]
        
        table_rows = [headers_def]
        
        for s in services:
            client_name = "N/A"
            if s.client:
                if hasattr(s.client, 'profile'):
                    client_name = f"{s.client.profile.first_name} {s.client.profile.last_name}".strip()
                if not client_name:
                    client_name = s.client.email
            if len(client_name) > 18:
                client_name = client_name[:16] + "..."
                
            worker_name = "N/A"
            if s.worker:
                if hasattr(s.worker, 'profile'):
                    worker_name = f"{s.worker.profile.first_name} {s.worker.profile.last_name}".strip()
                if not worker_name:
                    worker_name = s.worker.email
            if len(worker_name) > 18:
                worker_name = worker_name[:16] + "..."
                
            completed_str = s.completed_at.strftime('%Y-%m-%d') if s.completed_at else "N/A"
            price = float(s.total_price)
            comm = price * settings.PLATFORM_FEE_PERCENTAGE
            
            table_rows.append([
                Paragraph(str(s.id), table_cell_center),
                Paragraph(completed_str, table_cell_center),
                Paragraph(s.category.name if s.category else "N/A", table_cell_style),
                Paragraph(client_name, table_cell_style),
                Paragraph(worker_name, table_cell_style),
                Paragraph(f"${price:,.2f}", table_cell_right),
                Paragraph(f"${comm:,.2f}", table_cell_right)
            ])
            
        # Total Row in Table
        table_rows.append([
            Paragraph("TOTALES", table_header_style),
            Paragraph("", table_cell_style),
            Paragraph("", table_cell_style),
            Paragraph("", table_cell_style),
            Paragraph("", table_cell_style),
            Paragraph(f"${total_revenue:,.2f}", table_cell_right),
            Paragraph(f"${platform_commissions:,.2f}", table_cell_right)
        ])

        data_table = Table(table_rows, colWidths=[30, 65, 100, 115, 115, 60, 55])
        
        table_styles = [
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4F46E5')), # Indigo primary header
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('PADDING', (0,0), (-1,-1), 6),
            ('GRID', (0,0), (-1,-2), 0.5, colors.HexColor('#E2E8F0')),
            ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#CBD5E1')),
        ]
        
        # Apply zebra coloring to data rows (excluding header and total row)
        for i in range(1, len(table_rows) - 1):
            if i % 2 == 0:
                table_styles.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F8FAFC')))
                
        # Total row style
        total_row_idx = len(table_rows) - 1
        table_styles.extend([
            ('BACKGROUND', (0, total_row_idx), (-1, total_row_idx), colors.HexColor('#EEF2F6')),
            ('SPAN', (0, total_row_idx), (4, total_row_idx)),
            ('ALIGN', (0, total_row_idx), (4, total_row_idx), 'RIGHT'),
            ('TEXTCOLOR', (0, total_row_idx), (0, total_row_idx), colors.HexColor('#111827')),
            ('LINEABOVE', (0, total_row_idx), (-1, total_row_idx), 1.5, colors.HexColor('#4F46E5')),
        ])
        
        data_table.setStyle(TableStyle(table_styles))
        story.append(data_table)

        # Build Document
        doc.build(story)
        
        # Get PDF data and return
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_finanzas.pdf"'
        return response
